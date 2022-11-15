from openpyxl import load_workbook
import csv
from tkinter.filedialog import askopenfilename

filename = askopenfilename()
print(filename)

workbook = load_workbook(filename=filename)
print(workbook.sheetnames)
sheet = workbook.active
# sheets = workbook.sheetnames
# print(sheets)

sheet.title = "CIP_Pegs"
CIP_Bitmasks = workbook.create_sheet(title="CIP Bitmasks")

Header = ['N/A', 'Step Number', 'Step Description', 'Step Time', 'Max Step', 'Step Type', 'Validated',
          'Burst On', 'Burst Off', 'Bit Mask 1', 'Bit Mask 2', 'Bit Mask 3', 'Bit Mask 4',
          'Supply Flow', 'Wash Tank Level', 'Acid Tank Level', 'Recovery Tank Level', 'Rinse Tank Level',
          'Wash Conductivity', 'SP06', 'Supply Temp', 'SP08', 'SP09', 'SP10', 'SP11', 'SP12',
          'SP13', 'SP14', 'SP15', 'SP16', 'SP17', 'SP18', 'SP19', 'Hold for Supply Flow',
          'Hold for Wash Level', 'Hold for Wash Level', 'Hold for Acid Level', 'Hold for Recovery Level',
          'Hold for Rinse Level', 'HoldSP05', 'HoldSP06', 'Hold for Return Temp', 'Hold for Return Cond',
          'HoldSP09', 'HoldSP10', 'HoldSP11', 'HoldSP12', 'HoldSP13', 'HoldSP14', 'HoldSP15', 'HoldSP16',
          'HoldSP17', 'HoldSP18', 'HoldSP19']


Title = ['N/A', 'Step Number','Step Description','Step Time','Max Step Time','Step Type','Validated?',
         'Burst ON','Burst OFF','Bit Mask 1','Bit Mask 2','Bit Mask 3','Bit Mask 4','Supply Flow',
         'Wash Tank Level','Acid Tank Level','Recovery Tank Level','Rinse Tank Level','Wash Conductivity',
         'SP06','Supply Temp','SP08','SP09','SP10','SP11','SP12','SP13','SP14','S15','SP16','SP17','SP18',
         'SP19','Hold for Supply Flow','Hold for Wash Level','Hold for Acid Level','Hold for Recov Level',
         'Hold for Rinse Level','HoldSP05','HoldSP06','Hold Return Temp','Hold for Return Cond','HoldSP09',
         'HoldSP10','HoldSP11','HoldSP12','HoldSP13','HoldSP14','HoldSP15','HoldSP16','HoldSP17','HoldSP18','HoldSP19']

## Step type and description according to KPS Peg Editor
StepType = ['NA', 'Rinse', 'Wash', 'Acid', 'Recov', 'Sani']

StepDescription = ['NA', 'Ready Tanks', 'Fill CIP Tank', 'Charge Pre-Wash',
                   'Charge Caustic', 'Charge Acid', 'Establish Rinse', 'Establish Pre-Wash', 'Establish Caustic',
                   'Establish Acid', 'Establish Sani', 'Pre-Rinse', 'Rinse', 'Post-Rinse', 'Burst-Rinse', 'Pre-Wash',
                   'Caustic Wash', 'Acid Wash', 'Sani Wash', 'Sanitize', 'Recover Caustic', 'Recover Acid',
                   'Recover Rinse', 'Supply Air Blow', 'Air Blow', 'Pump Down', 'Drain', 'Drain to Sewer',
                   'Drain CIP Tank', 'Empty CIP Tank', 'Pump Out', 'Set Valves', 'Cool Down', 'Fat Save',
                   'Chlorinated Clean', 'Spray Rinse', 'Caustic Spray', 'Acid Spray', 'Lower Balance Tank',
                   'Flush Chemical', 'Add Chemical Press', 'Rinse from Recovery', 'Rinse Recirculate']
#print(StepDescription[sheet["B18"].value])

ABC_Bitmask = ['Full Wash', 'Sani Only', 'Rinse Only', 'Caustic Only', 'Acid Only', 'Sub Seq Enable', 'TOV Pulser',
               'Proof of Flow', 'Enable Flush Valves', 'Enable Drain Valves', 'Pump Down', 'PMO Valve Pulse Req',
               'Sani Pump', 'Chlor Tank Chem Pump', 'Acid Tank Chem Pump', 'Caustic Tank Chem Pump',
               'Force to Waste EQ Feed Tank', 'Return Pump', 'Supply Pump', 'Rinse Tank Outlet', 'Drain', 'Recovery Tank Outlet',
               'Recovery Tank Return', 'Chlor Tank Water', 'Chlor Tank Outlet', 'Chlor Tank Return', 'Sep Acid Tank Water',
               'Sep Acid Tank Outlet', 'Sep Acid Tank Return', 'Sep Caustic Tank Water', 'Sep Caustic Tank Outlet',
               'Sep Caustic Tank Return']
size_row = 50
size_column = 53

# Swap Description number for Step Description name
for row in range(1, size_row):
    print(sheet.cell(row=row, column=2).value)
    sheet.cell(row= row, column=2, value= StepDescription[sheet.cell(row= row, column=2).value])
    print(sheet.cell(row=row, column=2).value)

sheet.insert_rows(1)

for column in range(1,53):
    sheet.cell(row=1, column=column, value= Title[column])

## Bitmask Sheet----------------------------


for column in range(33):
    CIP_Bitmasks.cell(row=1, column=column, value=column)

## Save Workbook
workbook.save(filename=filename.replace('.xlsx', '_copy.xlsx'))

# Tgt_StepDes = 'B'
# Tgt_BM1 = 'I'
# Tgt_BM2 = 'J'
# #Tgt2 = 1
# Bitmasks1 = []
# Bitmasks2 = []
#
# for i in range(1,41):
#     Tgt = Tgt_StepDes + str(i)
#     sheet[Tgt].value = StepDescription[sheet[Tgt].value]
#
#     BM1_Tgt = Tgt_BM1 + str(i)
#     BM2_Tgt = Tgt_BM2 + str(i)
#
#     BM1 = bin(sheet[BM1_Tgt].value)
#     BM2 = bin(sheet[BM2_Tgt].value)
#     BM1 = BM1[2:]
#     BM2 = BM2[2:]
#
#     BM1 = [int(i) for i in BM1]
#     BM2 = [int(i) for i in BM2]
#
#     Bitmasks1.append(BM1)
#     Bitmasks2.append(BM2)
#
#
# workbook.save(filename=filename.replace('.xlsx', '_copy.xlsx'))
